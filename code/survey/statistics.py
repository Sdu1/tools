#!/usr/bin/python
# -*- coding: utf-8 -*-
import os
import xlrd
import xlwt
import numpy
import math

from openpyxl import load_workbook
from mysql import query, query_one, save


def get_file_path():
    all_path = os.walk('/mnt/hgfs/Data/work/survey/汇总/')

    xls_files = []
    xlsx_files = []
    for path in all_path:
        for file in path[-1]:
            if file.find(".xlsx") != -1:
                file = "%s/%s" % (path[0], file,)
                xlsx_files.append(file)
            elif file.find(".xls") != -1:
                file = "%s/%s" % (path[0], file,)
                xls_files.append(file)

    return (xls_files, xlsx_files,)


def extract_category(file_path):
    return unicode(file_path.split('/')[-1::1][0].split('.xls')[0], "utf-8")

def sheet_name_to_enterprise_name(sheet_name):
    enterprise_name = sheet_name.split(u"、")[1] if sheet_name.find(u"、") != -1 else sheet_name
    return enterprise_name

def enterprise_news_score(sheet_name):
    enterprise_name = sheet_name_to_enterprise_name(sheet_name)
    positive_count = query_one(sql=u'SELECT COUNT(*) FROM `base_news` WHERE `status` = 1 AND `keyword` LIKE "%%%s%%"', list1=(enterprise_name, ))
    negative_count = query_one(sql=u'SELECT COUNT(*) FROM `base_news` WHERE `status` = -1 AND `keyword` LIKE "%%%s%%"', list1=(enterprise_name, ))
    score = 7
    score += positive_count.get('COUNT(*)')
    score -= negative_count.get('COUNT(*)')

    if score < 0 :
        score = 0
    elif score > 10:
        score = 10

    return score

def average(score_list, sheet_name):
    avg = numpy.mean(score_list, axis=0)
    avg = round(numpy.mean([enterprise_news_score(sheet_name), avg], axis=0), 2) if not math.isnan(avg) else u'无评分'
    return avg

def handle_xls(xls_path):
    workbook = xlrd.open_workbook(xls_path)

    category = extract_category(xls_path)

    # get sheel
    sheet_names = filter(lambda x: x.find(
        u'heet') == -1, workbook.sheet_names())

    # accoding to sheel get sheel content
    data = {}
    for sheet_name in sheet_names:
        sheet = workbook.sheet_by_name(sheet_name)

        rownum = sheet.nrows  # row number
        colnum = sheet.ncols  # col number
        colnames = sheet.row_values(0)

        score_list = []
        for i in xrange(1, rownum):
            score_list += sheet.row_values(i)

        score_list = filter(
            lambda x: x != u'' and isinstance(x, float), score_list)
        data['%s : %s' % (category, sheet_name)] = average(score_list, sheet_name)

    return data


def handle_xlsx(xlsx_path):
    wb = load_workbook(filename=xlsx_path)

    category = extract_category(xlsx_path)

    sheet_names = filter(lambda x: x.find(u'heet') == -1,
                         wb.get_sheet_names())  # 获取所有表格(worksheet)的名字

    data = {}
    for sheet_name in sheet_names:
        ws = wb.get_sheet_by_name(sheet_name)  # 获取特定的 worksheet

        # 获取表格所有行和列，两者都是可迭代的
        rows = ws.rows
        columns = ws.columns

        # 行迭代
        score_list = []
        for index, row in enumerate(rows):
            if index > 0:
                for col in row:
                    x = col.value
                    if isinstance(x, long):
                        score_list.append(x)
        data['%s : %s' % (category, sheet_name)] = average(score_list, sheet_name)

    return data


def write_xls(data):
    file = xlwt.Workbook()                # 注意这里的Workbook首字母是大写
    table = file.add_sheet('statistics', cell_overwrite_ok=True)

    flag = 0
    for d in data:
        for k, v in d.items():
            flag += 1
            table.write(flag, 0, k)
            table.write(flag, 1, v)

    # 保存文件
    file.save('/home/sdu/Project/tools/code/survey/survey_statistics.xls')


def main():
    data = []
    xls_list = get_file_path()[0]
    for xls_path in xls_list:
        data.append(handle_xls(xls_path))

    xlsx_list = get_file_path()[1]
    for xlsx_path in xlsx_list:
        data.append(handle_xlsx(xlsx_path))

    write_xls(data)

if __name__ == '__main__':
    main()